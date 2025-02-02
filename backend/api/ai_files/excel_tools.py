#ne pas faire d'import circular
from openpyxl import load_workbook
from .classifier import theme_classifier, sentiment_classifier
import pandas as pd
import logging

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
import matplotlib.pyplot as plt
import geopandas as gpd
import os
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

def complete_excel(file_path):
    logging.info(file_path)
    # Charger le fichier Excel existant
    wb = load_workbook(file_path)
    ws = wb.active

    # Définir les en-têtes de colonnes attendus
    headers = ["Date", "Territoire", "Sujet", "Thème", "Qualité du retour", "Média", "Articles"]
    
    # Trouver l’index des colonnes en fonction des en-têtes
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]  # Première ligne
    header_indices = {header: i for i, header in enumerate(header_row) if header in headers}
    theme_qr = ["Thème", "Qualité du retour"]
    theme_qr_indices = {header: i for i, header in enumerate(header_row) if header in theme_qr}

    # Vérifier chaque ligne (à partir de la 2ème, car la 1ère est l’en-tête)
    i = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        i+=1
        logging.info(i/ws.max_row)
        for header, col_idx in theme_qr_indices.items():
            cell = row[col_idx]  # Récupérer la cellule correspondant à la colonne
            if cell.value is None or str(cell.value).strip() == "":
                #cell.value = "1"  # Remplacer les valeurs vides par None
                if headers[col_idx] == "Thème":
                    article = {}
                    article["Date"] = str(row[header_indices['Date']].value).strip()
                    article["Territoire"] = str(row[header_indices['Territoire']].value).strip()
                    article["Sujet"] = str(row[header_indices['Sujet']].value).strip()
                    article["Média"] = str(row[header_indices['Média']].value).strip()
                    article["Articles"] = str(row[header_indices['Articles']].value).strip()
                    cell.value = theme_classifier(article)['Thème']
                if headers[col_idx] == "Qualité du retour":
                    cell.value = sentiment_classifier(str(row[header_indices['Articles']].value).strip())

    # Sauvegarder les modifications
    wb.save(file_path)


############################################################################################################
import reportlab
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
import matplotlib.pyplot as plt
import numpy as np
import geopandas as gpd
import pandas as pd
import os
from datetime import datetime
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import Paragraph
from openpyxl import load_workbook

import re
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import Paragraph
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm

SLIDE_WIDTH = 1.5*17 * cm  # Largeur type PowerPoint en cm
SLIDE_HEIGHT = 1.5*13 * cm  # Hauteur type PowerPoint en cm

import time
import pandas as pd
import boto3
import os
import re
import json
import time
import logging
import base64
from concurrent.futures import ThreadPoolExecutor

def llm_call(message, image=None):
    temperature = 1
    max_tokens = 8192
    region = 'us-west-2'
    model_id = "anthropic.claude-3-5-sonnet-20241022-v2:0"


    if image:
        with open(image, "rb") as image_file:
            content_image = base64.b64encode(image_file.read()).decode('utf8')
        messages = {"role": "user",
                "content": [
                {"type": "image", "source": {"type": "base64",
                    "media_type": "image/jpeg", "data": content_image}},
                {"type": "text", "text": message}
                ]}
    else:
         messages = {"role": "user",
                "content": [
                {"type": "text", "text": message}
                ]}
    
    body = json.dumps(
        {
            "anthropic_version": "bedrock-2023-05-31",
            "max_tokens": max_tokens,
             "messages": [messages]
        }
    )

    bedrock_client = boto3.client(
        'bedrock-runtime',  # Utilisation du service AWS Bedrock
        region_name=region # À ajuster selon la région AWS que vous utilisez
    )    
    
    try:
        # Appeler l'API Bedrock
        response = bedrock_client.invoke_model(body=body, modelId=model_id)
        response_body = json.loads(response.get('body').read())

        return response_body['content'][0]['text']

    except Exception as e:
        print(f"Erreur lors de l'appel à l'API Bedrock: {e}")

def get_comments(comment_prompt, image_path):
    comments = llm_call(comment_prompt, image_path)
    matches = re.findall(r"\[([^\]]+)\]", comments)

    # Convertir en bullet points avec sauts de ligne corrects
    formatted_comments = "• " + "<br/>• ".join(matches)

    return formatted_comments

def add_slide_with_comment(c, image_path, formatted_comments, y_position=14):
    """
    Ajoute une slide avec une image et des commentaires bien formatés.
    
    :param c: Canvas du PDF
    :param image_path: Chemin de l'image à insérer
    :param comment_prompt: Prompt pour récupérer les commentaires via llm_call
    :param y_position: Hauteur de l'image en cm (ajuster si besoin)
    """
    width, height = SLIDE_WIDTH, SLIDE_HEIGHT

    # Ajuster les marges et les positions
    image_width = 15 * cm
    image_height = 10 * cm  # Réduire légèrement si besoin
    image_x = (width - image_width) / 2  # Centrer l'image
    image_y = height - y_position * cm  # Ajuster la hauteur

    # Insérer l'image centrée
    c.drawImage(image_path, image_x, image_y, width=image_width, height=image_height)


    # Définir un style de paragraphe
    style = ParagraphStyle(name="CommentStyle", fontSize=12, leading=14)

    # Déterminer l'emplacement du texte sous l’image
    text_x = 2 * cm
    text_y = image_y - 2.5 * cm  # Ajuster la hauteur sous l'image

    # Créer et afficher le paragraphe
    comment_paragraph = Paragraph(formatted_comments, style)
    text_width = width - 4 * cm  # Garde des marges sur les côtés
    text_height = text_y - 2 * cm  # Ajuste la hauteur max pour ne pas dépasser

    comment_paragraph.wrapOn(c, text_width, text_height)
    comment_paragraph.drawOn(c, text_x, text_y)

    # Nouvelle page après chaque slide
    c.showPage()


def draw_summary(c, summary_text, x=2*cm, y=20*cm, width=16*cm):
    # Définir le style du texte
    style = ParagraphStyle(name="SummaryStyle", fontSize=14, leading=16)

    
    # Créer un paragraphe avec le texte formaté
    paragraph = Paragraph(summary_text, style)
    
    # Afficher le texte dans le PDF
    paragraph.wrapOn(c, width, 20*cm)  # Largeur max et hauteur dispo
    paragraph.drawOn(c, x, y)
    c.showPage()

def generate_pdf_from_dataframe(df, pdf_filename="rapport_graphes.pdf"):
    
    geojson_path = "./api/ai_files/files_pdfgenerator/departements.geojson"
    enedis_logo_path = "./api/ai_files/files_pdfgenerator/enedis_logo.png"
    ia_logo_path = "./api/ai_files/files_pdfgenerator/ia.png"
    
    width, height = SLIDE_WIDTH, SLIDE_HEIGHT
    temp_path = f"result_{pdf_filename}"
    c = canvas.Canvas(temp_path, pagesize=(SLIDE_WIDTH, SLIDE_HEIGHT))
    comments = []

    c.drawImage(enedis_logo_path, 3 * cm, height - 10 * cm, width=15 * cm, height=12 * cm)
    date_formatee = datetime.now().strftime("%d %B %Y")
    c.setFont("Helvetica", 18)
    c.drawCentredString(width / 2, height - 7 * cm, date_formatee)
    c.drawImage(ia_logo_path, 1 * cm, height - 12 * cm, width=5 * cm, height=3 * cm)
    c.showPage()

    c.setFont("Helvetica-Bold", 22)
    c.drawString(2 * cm, height - 3 * cm, "Sommaire")
    c.setFont("Helvetica", 14)
    c.drawString(3 * cm, height - 6 * cm, "Résumé ....................................... Page 3")
    c.drawString(3 * cm, height - 8 * cm, "Graphiques .................. Page 4")
    c.showPage()
    c.setFont("Helvetica-Bold", 22)
    c.drawString(2 * cm, height - 3 * cm, "Résumé")

    graph_bar_theme_filename = "graph_bar_quality_theme.png"
    graph_pos_filename = "graph_positive_pie.png"
    graph_neg_filename = "graph_negative_pie.png"
    graph_bar_filename = "graph_bar_quality.png"
    graph_time_filename = "graph_time.png"

    graphs = [
        (graph_pos_filename, "Ce graphe représente la distribution des tonalités par média en terme d'image d'Enedis qui est transmise dans leurs articles. Donne les 2 commentaires les plus pertinents à ajouter sur la slide avec le graphe. Cela doit être très consis (en quelques mots). Tu retournera entre crochets ce qui doit apparaître sur les slides"),
        (graph_neg_filename, "Ce graphe représente la répartition des avis négatifs par média sur Enedis. Donne les 2 commentaires les plus pertinents à ajouter sur la slide avec le graphe. Cela doit être très consis (en quelques mots). Tu retournera entre crochets ce qui doit apparaître sur les slides"),
        (graph_bar_filename, "Voici une visualisation des proportions des retours (positif, négatif, factuel) par média. Donne les 2 commentaires les plus pertinents à ajouter sur la slide. Cela doit être très consis (en quelques mots). Tu retournera entre crochets ce qui doit apparaître sur les slides"),
        (graph_bar_theme_filename, "Proportions des retours par Thème. Quels enseignements peut-on tirer de cette analyse ? Donne les 2 commentaires les plus pertinents. Cela doit être très consis (en quelques mots). Tu retournera entre crochets ce qui doit apparaître sur les slides"),
        (graph_time_filename, "Évolution des retours sur Enedis dans le temps. Quels sont les éléments clés qui se dégagent ? Donne les 2 commentaires les plus pertinents. Cela doit être très consis (en quelques mots). Tu retournera entre crochets ce qui doit apparaître sur les slides"),
    ]

    df_positive = df[df["Qualité du retour"] == "positif"]
    df_negative = df[df["Qualité du retour"] == "négatif"]
    df_factuel = df[df["Qualité du retour"] == "factuel"]

    # Camembert des retours positifs
    fig, ax = plt.subplots(figsize=(6, 6))
    df_positive["Média"].value_counts().plot.pie(autopct='%1.1f%%', startangle=90, colormap="Greens", ax=ax)
    ax.set_ylabel("")
    ax.set_title("Répartition des retours positifs")
    
    fig.savefig(graph_pos_filename, bbox_inches="tight", dpi=150)
    plt.close(fig)

    # Camembert des retours négatifs
    fig, ax = plt.subplots(figsize=(6, 6))
    df_negative["Média"].value_counts().plot.pie(autopct='%1.1f%%', startangle=90, colormap="Reds", ax=ax)
    ax.set_ylabel("")
    ax.set_title("Répartition des retours négatifs")
    graph_neg_filename = "graph_negative_pie.png"
    fig.savefig(graph_neg_filename, bbox_inches="tight", dpi=150)
    plt.close(fig)

    # Graphique en barres empilées
    media_counts = df.groupby(["Média", "Qualité du retour"]).size().unstack(fill_value=0)
    media_counts = media_counts.div(media_counts.sum(axis=1), axis=0).iloc[1:]  # Normalisation

    fig, ax = plt.subplots(figsize=(8, 6))
    media_counts.plot(kind='bar', stacked=True, ax=ax, color=['red', 'green', 'gray'])
    ax.set_title("Proportions des retours par Média")
    ax.set_xlabel("Média")
    ax.set_ylabel("Proportion")
    ax.legend(["Négatif", "Positif", "Factuel"], loc="best", fontsize=8)
    plt.xticks(rotation=45)
    graph_bar_filename = "graph_bar_quality.png"
    fig.savefig(graph_bar_filename, bbox_inches="tight", dpi=150)
    plt.close(fig)

    # Graphique temporel
    df_time = df.groupby(["Date", "Qualité du retour"]).size().unstack(fill_value=0).iloc[:-1]
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.bar(df_time.index, df_time.get("positif", 0), color='green', label='Positif')
    ax.bar(df_time.index, -df_time.get("négatif", 0), color='red', label='Négatif')
    ax.set_title("Évolution des retours dans le temps")
    ax.set_xlabel("Date")
    ax.set_ylabel("Nombre de retours")
    ax.legend()
    map_filename = "map_france.png"
    graph_time_filename = "graph_time.png"
    fig.savefig(graph_time_filename, bbox_inches="tight", dpi=150)
    plt.close(fig)

    # Graphique en barres par thème
    media_counts = df.groupby(["Thème", "Qualité du retour"]).size().unstack(fill_value=0)
    media_counts = media_counts.div(media_counts.sum(axis=1), axis=0).iloc[1:]  # Normalisation

    fig, ax = plt.subplots(figsize=(8, 6))
    media_counts.plot(kind='bar', stacked=True, ax=ax, color=['red', 'green', 'gray'])
    ax.set_title("Proportions des retours par Thème")
    ax.set_xlabel("Thème")
    ax.set_ylabel("Proportion")
    ax.legend(["Négatif", "Positif", "Factuel"], loc="best", fontsize=8)
    plt.xticks(rotation=45)

    
    fig.savefig(graph_bar_theme_filename, bbox_inches="tight", dpi=150)
    plt.close(fig)

    comments = []
    resume_context = "Fais moi un résumé des mes slides graces aux commentaires des graphs suivant générés :"

    for file, prompt in graphs:
        comment = get_comments(prompt,file)
        comments.append(comment)
        logging.info(comment)
        resume_context += "\n\n" + prompt + "\n\n" + comment
    
    resume_context += "\n\n Tu donnera ton résumé entre crochet."
    response = llm_call(resume_context, None)
    matches = re.findall(r"\[([^\]]+)\]", response)
    # Convertir en bullet points avec sauts de ligne corrects

    resume = "\n".join(matches)
    draw_summary(c, resume)
    logging.info(resume)

    # Charger la carte de la France depuis un GeoJSON
    france = gpd.read_file(geojson_path)
    
    # Vérification du système de coordonnées et reprojection si nécessaire
    if france.crs is None or france.crs.to_epsg() != 3857:
        france = france.to_crs(epsg=3857)
    
    # Comptage des retours positifs et négatifs par département
    df["Territoire"] = df["Territoire"].str.lower()
    dept_counts = df.groupby(["Territoire", "Qualité du retour"]).size().unstack(fill_value=0)
    dept_counts["color"] = dept_counts.apply(lambda x: "red" if x.get("négatif", 0) > x.get("positif", 0) else ("green" if x.get("positif", 0) > x.get("négatif", 0) else "gray"), axis=1)
    
    # Fusionner avec les données géographiques
    france["nom"] = france["nom"].str.lower()
    france = france.merge(dept_counts, how="left", left_on="nom", right_index=True)
    france = france.dropna(subset=["color"])  # Supprimer les départements sans valeurs associées
        
    # Création de la figure avec fond de carte
    fig, ax = plt.subplots(figsize=(6, 6))
    france.plot(ax=ax, color=france["color"].fillna("white"), edgecolor="black")
    
    ax.set_xticks([])
    ax.set_yticks([])
    ax.set_frame_on(False)
    
    # Sauvegarde temporaire de la figure
    map_filename = "map_france.png"
    fig.savefig(map_filename, bbox_inches="tight", dpi=150)
    plt.close(fig)

    c.setFont("Helvetica-Bold", 22)
    c.drawString(2 * cm, height - 3 * cm, "Graphiques")
    c.setFont("Helvetica-Bold", 20)
    c.drawString(2 * cm, height - 5 * cm, "1. Géographie")
    
    # Insérer la carte dans le PDF
    c.drawImage(map_filename, 2 * cm, height - 18 * cm, width=15 * cm, height=12 * cm)
    
    c.showPage()
    os.remove(map_filename)

    c.setFont("Helvetica-Bold", 20)
    c.drawString(2 * cm, height - 3 * cm, "2. Analyse des articles par Média")
    

    # Stockage des images à afficher
   


    for i, (file, prompt) in enumerate(graphs):
        if file == graph_time_filename:
            c.setFont("Helvetica-Bold", 20)
            c.drawString(2 * cm, height - 4 * cm, "3. Graphique temporel")
        add_slide_with_comment(c, file, comments[i])

    c.save()

    os.remove(graph_pos_filename)
    os.remove(graph_neg_filename)
    os.remove(graph_bar_filename)
    os.remove(graph_time_filename)

    return temp_path

def generate_pdf_result(excel_path_file):
    logging.info(excel_path_file)
    # Charger le fichier Excel
    wb = load_workbook(excel_path_file)
    ws = wb.active

    headers = ["Date", "Territoire", "Sujet", "Thème", "Qualité du retour", "Média", "Articles"]
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]  
    header_indices = {header: i for i, header in enumerate(header_row) if header in headers}

    data = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        data.append({header: row[col_idx] for header, col_idx in header_indices.items()})
    df = pd.DataFrame(data)

    df.columns = df.columns.astype(str)
    df.columns = df.columns.str.strip()
    df.columns = df.columns.str.replace("\n", " ")

    temp_pdf_path = generate_pdf_from_dataframe(df)
    return temp_pdf_path



